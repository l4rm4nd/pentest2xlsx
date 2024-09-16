#!/usr/bin/env python3
#    Copyright (C) 2019 Alexandre Teyar

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

# http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
#    limitations under the License.

# TODO:
# * add a "File" column

from . import config
from ..parser import Parser
from pathlib import Path

import json
import logging
import re
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

class Testssl(Parser):
    def __init__(self, input_files, output_file):
        super(Testssl, self).__init__(input_files, output_file)

    def print_vars(self):
        logging.info("input file(s): {}".format(
            sorted([x.name for x in self._input_files])))
        logging.info("output file: {}".format(self._output_file))
        logging.info("certificate issue(s) to process: {}".format(
            sorted(config.certificates.keys())))
        logging.info("protocol(s) to process: {}".format(
            sorted(config.protocols)))
        logging.info("vulnerability/ies to process: {}".format(
            sorted(config.vulnerabilities.keys())))
        print()

    def parse(self):
        logging.info("generating worksheet 'Host vs Certificate (Details)'...")
        self.parse_host_certificate()
        print()
        logging.info("generating worksheet 'Host vs Certificates'...")
        self.parse_host_certificates()
        print()
        logging.info("generating worksheet 'Host vs Protocol (Details)'...")
        self.parse_host_protocol()
        print()
        logging.info("generating worksheet 'Host vs Protocols'...")
        self.parse_host_protocols()
        print()
        logging.info("generating worksheet 'Host vs Vulnerability (Details)'...")
        self.parse_host_vulnerability()
        print()
        logging.info("generating worksheet 'Host vs Vulnerabilities'...")
        self.parse_host_vulnerabilities()
        print()

        try:
            self._workbook.close()
            
            try:
                logging.info("removing duplicates...")
                self.remove_duplicates()
            except:
                pass
            logging.info("formatting worksheets...")
            self.apply_formatting()
            print()
        except Exception as e:
            logging.exception("{}".format(e))

    def remove_duplicates(self):
        # Load the workbook
        workbook = load_workbook(self._output_file)

        # Iterate through each sheet in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Read the sheet data into a DataFrame
            df = pd.read_excel(self._output_file, sheet_name=sheet_name)

            # Remove duplicates while ignoring the "File" column
            df = df.drop_duplicates(subset=df.columns.difference(['File']))

            # Write the updated DataFrame back to the sheet
            with pd.ExcelWriter(self._output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                writer.book = workbook
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Save the workbook with the updated sheets
        workbook.save(self._output_file)

    def apply_formatting(self):
        # Load the workbook
        workbook = load_workbook(self._output_file)

        # Define the conditional formatting rules
        rules = [
            ("OK", "00FF00"),         # Green background for "OK"
            ("LOW", "FFFF00"),        # Yellow background for "LOW"
            ("MEDIUM", "FFA500"),     # Orange background for "MEDIUM"
            ("HIGH", "FF0000"),       # Red background for "HIGH"
            ("CRITICAL", "8B0000"),   # Dark red background for "CRITICAL"
            ("INFO", "D3D3D3")        # Light grey background for "INFO"
        ]

        # Iterate through each sheet in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Apply conditional formatting for the rules
            for rule in rules:
                search_text, fill_color = rule
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and search_text in str(cell.value):
                            cell.fill = fill

        # Check if "Host vs Protocols" tab exists in the workbook
        if "Host vs Protocols" in workbook.sheetnames:
            sheet = workbook["Host vs Protocols"]

            # Get the column indexes for SSLv2, SSLv3, and TLS1
            sslv2_index = None
            sslv3_index = None
            tls1_index = None

            for cell in sheet[1]:
                column_name = cell.value
                if column_name == "SSLv2":
                    sslv2_index = cell.column
                elif column_name == "SSLv3":
                    sslv3_index = cell.column
                elif column_name == "TLS1":
                    tls1_index = cell.column
                elif column_name == "TLS1_1":
                    tls11_index = cell.column
                elif column_name == "TLS1_2":
                    tls12_index = cell.column
                elif column_name == "TLS1_3":
                    tls13_index = cell.column

            # Apply formatting based on column values
            if sslv2_index or sslv3_index or tls1_index or tls11_index or tls12_index or tls13_index:
                for row in sheet.iter_rows(min_row=2):
                    if sslv2_index:
                        sslv2_cell = row[sslv2_index - 1]
                        sslv2_value = str(sslv2_cell.value).strip().lower()
                        sslv2_cell.fill = PatternFill(start_color="FF0000" if sslv2_value == "yes" else "00FF00",
                                                      end_color="FF0000" if sslv2_value == "YES" else "00FF00",
                                                      fill_type="solid")
                    if sslv3_index:
                        sslv3_cell = row[sslv3_index - 1]
                        sslv3_value = str(sslv3_cell.value).strip().lower()
                        sslv3_cell.fill = PatternFill(start_color="FF0000" if sslv3_value == "yes" else "00FF00",
                                                      end_color="FF0000" if sslv3_value == "YES" else "00FF00",
                                                      fill_type="solid")
                    if tls1_index:
                        tls1_cell = row[tls1_index - 1]
                        tls1_value = str(tls1_cell.value).strip().lower()
                        tls1_cell.fill = PatternFill(start_color="FFFF00" if tls1_value == "yes" else "00FF00",
                                                      end_color="FFFF00" if tls1_value == "YES" else "00FF00",
                                                      fill_type="solid")
                    if tls11_index:
                        tls11_cell = row[tls11_index - 1]
                        tls11_value = str(tls11_cell.value).strip().lower()
                        tls11_cell.fill = PatternFill(start_color="FFFF00" if tls11_value == "yes" else "00FF00",
                                                      end_color="FFFF00" if tls11_value == "YES" else "00FF00",
                                                      fill_type="solid")
                    if tls12_index:
                        tls12_cell = row[tls12_index - 1]
                        tls12_value = str(tls12_cell.value).strip().lower()
                        tls12_cell.fill = PatternFill(start_color="00FF00" if tls12_value == "yes" else "FFFF00",
                                                      end_color="00FF00" if tls12_value == "YES" else "FFFF00",
                                                      fill_type="solid")
                    if tls13_index:
                        tls13_cell = row[tls13_index - 1]
                        tls13_value = str(tls13_cell.value).strip().lower()
                        tls13_cell.fill = PatternFill(start_color="00FF00" if tls13_value == "yes" else "FFFF00",
                                                      end_color="00FF00" if tls13_value == "YES" else "FFFF00",
                                                      fill_type="solid")     
        # Iterate through each sheet in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Autofit columns for the sheet
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                # Find the maximum length of data in the column
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass

                # Set the column width to fit the maximum length
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Save the modified workbook
        workbook.save(self._output_file)
        workbook.close()

    def parse_host_certificate(self):
        table_data = []
        table_headers = [
            {"header": "File"},
            {"header": "Host IP"},
            {"header": "Host DNS"},
            {"header": "Port"},
            {"header": "Vulnerability"},
            {"header": "Severity"},
            {"header": "Information"}
        ]

        for input_file in self._input_files:
            input_file.seek(0)
            host_certificates = get_host_certificates(input_file)

            for values in host_certificates:
                for certificates in values["certificates"]:
                    table_data.append(
                        [
                            values["filepath"],
                            values["host_ip"],
                            values["host_dns"],
                            int(values["port"]),
                            certificates["vulnerability"],
                            certificates["severity"],
                            certificates["information"]
                        ]
                    )

        worksheet = self._workbook.add_worksheet("Host vs Certificate (Details)")
        self.draw_table(worksheet, table_headers, table_data)

    def parse_host_certificates(self):
        table_data = []
        table_headers = [
            {"header": "File"},
            {"header": "Host IP"},
            {"header": "Host DNS"},
            {"header": "Port"}
        ]

        # Add certificate names as headers from config.certificates
        for values in config.certificates.values():
            table_headers.append({"header": values["name"]})

        try:
            for input_file in self._input_files:
                input_file.seek(0)
                host_certificates = self.get_host_certificates_for_tab(input_file)  # Call the method via 'self'

                for values in host_certificates:
                    table_data.append(reindexHeaders(table_headers, values))

            worksheet = self._workbook.add_worksheet("Host vs Certificates")
            self.draw_table(worksheet, table_headers, table_data)
        except KeyError as e:
            logging.exception("KeyError: {}".format(e))
        except ValueError as e:
            logging.exception("ValueError: {}".format(e))

    def get_host_certificates_for_tab(self, file):
        results = []

        try:
            data = json.load(file)

            for values in data["scanResult"]:
                certificates = {}

                # Add the file path, host IP, DNS, and port info to the results dictionary
                certificates["filepath"] = Path(file.name).resolve().as_posix()
                certificates["host_ip"] = values["ip"]
                certificates["host_dns"] = values["targetHost"]
                certificates["port"] = int(values["port"])

                # Iterate over all configured certificate keys
                for cert_key, cert_value in config.certificates.items():
                    found = False
                    cert_name = cert_value["name"]

                    # Check for both regular and <hostCert#1> cert_key
                    for serverDefault in values.get("serverDefaults", []):
                        # If the regular cert_key (e.g., cert_chain_of_trust) exists
                        if serverDefault["id"] == cert_key:
                            certificates[cert_name] = serverDefault["severity"]
                            found = True
                            break  # Stop once the regular key is found

                        # If <hostCert#1> exists and no regular key has been found, treat it as regular
                        elif re.match(rf"^{cert_key}\s*<hostCert#1>$", serverDefault["id"]):
                            certificates[cert_name] = serverDefault["severity"]
                            found = True
                            break  # Stop once <hostCert#1> is found

                results.append(certificates)

        except KeyError as e:
            logging.exception("KeyError: {}".format(e))
            logging.exception("Skipping JSON file: " + file.name)
        except ValueError as e:
            logging.exception("ValueError: {}".format(e))
            logging.exception("Skipping JSON file: " + file.name)

        return results

    def parse_host_protocol(self):
        table_data = []
        table_headers = [
            {"header": "File"},
            {"header": "Host IP"},
            {"header": "Host DNS"},
            {"header": "Port"},
            {"header": "Supported Protocol"},
            {"header": "Severity"},
            {"header": "Information"}
        ]

        for input_file in self._input_files:
            input_file.seek(0)
            host_protocols = get_host_protocols(input_file)

            for values in host_protocols:
                for protocol in values["protocols"]:
                    table_data.append(
                        [
                            values["filepath"],
                            values["host_ip"],
                            values["host_dns"],
                            int(values["port"]),
                            protocol["version"],
                            protocol["severity"],
                            protocol["finding"]
                        ]
                    )

        worksheet = self._workbook.add_worksheet("Host vs Protocol (Details)")
        self.draw_table(worksheet, table_headers, table_data)

    def parse_host_protocols(self):
        table_data = []
        table_headers = [
            {"header": "File"},
            {"header": "Host IP"},
            {"header": "Host DNS"},
            {"header": "Port"}
        ]

        for protocol in config.protocols:
            table_headers.append({"header": protocol})

        try:
            for input_file in self._input_files:
                input_file.seek(0)
                host_protocols = get_host_protocols(input_file)

                for values in host_protocols:
                    table_data.append(reindex(table_headers, values))

            worksheet = self._workbook.add_worksheet("Host vs Protocols")
            self.draw_table(worksheet, table_headers, table_data)
        except KeyError as e:
            logging.exception("KeyError: {}".format(e))
        except ValueError as e:
            logging.exception("ValueError: {}".format(e))

    def parse_host_vulnerability(self):
        table_data = []
        table_headers = [
            {"header": "File"},
            {"header": "Host IP"},
            {"header": "Host DNS"},
            {"header": "Port"},
            {"header": "Vulnerability"},
            {"header": "Severity"},
            {"header": "CVE"},
            {"header": "Information"}
        ]

        for input_file in self._input_files:
            input_file.seek(0)
            host_vulnerabilities = get_host_vulnerabilities(input_file)

            for values in host_vulnerabilities:
                for vulnerability in values["vulnerabilities"]:
                    table_data.append(
                        [
                            values["filepath"],
                            values["host_ip"],
                            values["host_dns"],
                            int(values["port"]),
                            vulnerability["vulnerability"],
                            vulnerability["severity"],
                            vulnerability["cve"],
                            vulnerability["information"]
                        ]
                    )

        worksheet = self._workbook.add_worksheet("Host vs Vulnerability (Details)")
        self.draw_table(worksheet, table_headers, table_data)

    def parse_host_vulnerabilities(self):
        table_data = []
        table_headers = [
            {"header": "File"},
            {"header": "Host IP"},
            {"header": "Host DNS"},
            {"header": "Port"}
        ]

        for values in config.vulnerabilities.values():
            table_headers.append({"header": values["name"]})

        try:
            for input_file in self._input_files:
                input_file.seek(0)
                host_vulnerabilities = get_host_vulnerabilities(input_file)

                for values in host_vulnerabilities:
                    table_data.append(reindex(table_headers, values))

            worksheet = self._workbook.add_worksheet("Host vs Vulnerabilities")
            self.draw_table(worksheet, table_headers, table_data)
        except KeyError as e:
            logging.exception("KeyError: {}".format(e))
        except ValueError as e:
            logging.exception("ValueError: {}".format(e))


def reindex(table_headers, d):
    results = [None for y in range(len(table_headers))]

    results[table_headers.index({'header': 'File'})] = d["filepath"]
    results[table_headers.index({'header': 'Host IP'})] = d["host_ip"]
    results[table_headers.index({'header': 'Host DNS'})] = d["host_dns"]
    results[table_headers.index({'header': 'Port'})] = d["port"]

    for key, values in d.items():
        if isinstance(values, list):
            for value in values:
                try:
                    if key in {"certificates", "vulnerabilities"}:
                        results[table_headers.index(
                            {'header': value["vulnerability"]})] = \
                            value["severity"]
                    elif key == "protocols":
                        results[table_headers.index(
                            {'header': value["version"]})] = \
                            value["is_offered"]
                except ValueError:
                    logging.exception("code needs to be inspected")

    return results

def reindexHeaders(table_headers, d):
    results = [None for _ in range(len(table_headers))]

    # Map basic fields: File, Host IP, Host DNS, and Port
    results[table_headers.index({"header": "File"})] = d.get("filepath", "")
    results[table_headers.index({"header": "Host IP"})] = d.get("host_ip", "")
    results[table_headers.index({"header": "Host DNS"})] = d.get("host_dns", "")
    results[table_headers.index({"header": "Port"})] = d.get("port", "")

    # Map certificates dynamically based on config
    # Handle <hostCert#X> by normalizing them to standard headers
    for header in table_headers[4:]:  # Skip the first four headers (File, IP, DNS, Port)
        cert_name = header["header"]  # This is the certificate name, e.g., "Chain of Trust"
        matching_key = None

        # Try to find the matching certificate in the data, handling <hostCert#X>
        for key in d.keys():
            # Normalize keys like cert_chain_of_trust <hostCert#1> to cert_chain_of_trust
            if key == cert_name or re.match(rf"{cert_name} <hostCert#\d+>", key):
                matching_key = key
                break

        if matching_key:
            results[table_headers.index({"header": cert_name})] = d[matching_key]

    return results    


def get_host_certificates(file):
    results = []

    try:
        data = json.load(file)

        for values in data["scanResult"]:
            certificates = []

            # Iterate over all config.certificates keys
            for cert_key, cert_value in config.certificates.items():
                # First, check for the standard cert_key (e.g., "cert_chain_of_trust")
                for serverDefault in values["serverDefaults"]:
                    if serverDefault["id"] == cert_key:
                        certificates.append({
                            "vulnerability": cert_value["name"],
                            "severity": serverDefault["severity"],
                            "information": serverDefault["finding"]
                        })

                # Now, look for dynamically named certs like "cert_key <hostCert#X>"
                cert_pattern = re.compile(rf"^{cert_key}\s*<hostCert#1+>$")
                for serverDefault in values["serverDefaults"]:
                    if cert_pattern.match(serverDefault["id"]):
                        certificates.append({
                            "vulnerability": cert_value["name"],
                            "severity": serverDefault["severity"],
                            "information": serverDefault["finding"]
                        })

            # Add the parsed certificates to the results
            results.append({
                "filepath": Path(file.name).resolve().as_posix(),
                "host_ip": values["ip"],
                "host_dns": values["targetHost"],
                "port": int(values["port"]),
                "certificates": certificates
            })

    except KeyError as e:
        logging.exception("KeyError: {}".format(e))
        logging.exception("Skipping JSON file: " + file.name)
    except ValueError as e:
        logging.exception("ValueError: {}".format(e))
        logging.exception("Skipping JSON file: " + file.name)

    return results

def get_host_protocols(file):
    results = []

    try:
        data = json.load(file)

        for values in data["scanResult"]:
            protocols = []

            for protocol in values["protocols"]:
                if protocol["id"] in config.protocols:
                    if "not offered" not in protocol["finding"]:
                        protocols.append(
                            {
                                "version": protocol["id"],
                                "is_offered": "YES",
                                "severity": protocol["severity"],
                                "finding": protocol["finding"]
                            }
                        )
                    else:
                        protocols.append(
                            {
                                "version": protocol["id"],
                                "is_offered": "NO",
                                "severity": protocol["severity"],
                                "finding": protocol["finding"]
                            }
                        )

            results.append(
                {
                    "filepath": Path(file.name).resolve().as_posix(),
                    "host_ip": values["ip"],
                    "host_dns": values["targetHost"],
                    "port": int(values["port"]),
                    "protocols": protocols
                }
            )
    except KeyError as e:
        logging.exception("KeyError: {}".format(e))
        logging.exception("Skipping JSON file: " + file.name)
    except ValueError as e:
        logging.exception("ValueError: {}".format(e))
        logging.exception("Skipping JSON file: " + file.name)

    return results


def get_host_vulnerabilities(file):
    results = []

    try:
        data = json.load(file)

        for values in data["scanResult"]:
            vulnerabilities = []

            for vulnerability in values["vulnerabilities"]:
                if vulnerability["id"] in config.vulnerabilities.keys():
                    vulnerabilities.append(
                        {
                            "vulnerability": config.
                            vulnerabilities[vulnerability["id"]]["name"],
                            "severity": vulnerability["severity"],
                            "cve": vulnerability.get("cve", "N/A")
                            .replace(" ", ";"),
                            "information": vulnerability["finding"]
                        }
                    )

            for vulnerability in values["ciphers"]:
                if vulnerability["id"] in config.vulnerabilities.keys():
                    vulnerabilities.append(
                        {
                            "vulnerability": config.
                            vulnerabilities[vulnerability["id"]]["name"],
                            "severity": vulnerability["severity"],
                            "cve": vulnerability.get("cwe", "N/A")
                            .replace(" ", ";"),
                            "information": vulnerability["finding"]
                        }
                    )

            results.append(
                {
                    "filepath": Path(file.name).resolve().as_posix(),
                    "host_ip": values["ip"],
                    "host_dns": values["targetHost"],
                    "port": int(values["port"]),
                    "vulnerabilities": vulnerabilities
                }
            )
    except KeyError as e:
        logging.exception("KeyError: {}".format(e))
        logging.exception("Skipping JSON file: " + file.name)
        
    except ValueError as e:
        logging.exception("ValueError: {}".format(e))
        logging.exception("Skipping JSON file: " + file.name)

    return results
